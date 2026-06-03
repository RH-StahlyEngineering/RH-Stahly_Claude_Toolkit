"""Canonical resource resolver.

Reads ``references/stahly_canonical_paths.md``, verifies the cached
``resolved_path`` against fingerprint markers, falls back to ``search_hints``
glob walks if the cached path is gone or stale, and writes the new path back
on success. **Refuses to silently fall back on miss** — raises
``CanonicalResourceNotFound`` so the orchestrator escalates to the user.

Public API
----------
``resolve(resource_id, *, force_refresh=False) -> Path``
    Return the resolved path for the resource, or raise.

``resolve_all() -> dict``
    Resolve every resource in the registry. Returns ``{resource_id: Path}``.
    Raises on the first miss.

``update_resolved_path(resource_id, new_path) -> None``
    Write a new ``resolved_path`` into the registry file + log it.

``CanonicalResourceNotFound``
    Structured exception carrying ``resource_id``, ``tried_hints``,
    ``candidates_seen``, ``failed_markers``. The orchestrator catches this and
    presents a clear question to the user.
"""
from __future__ import annotations

import datetime as _dt
import fnmatch
import glob
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

# Force UTF-8 on Windows so non-ASCII status glyphs (✗ ✓ ⚠ → — etc.) don't
# crash with UnicodeEncodeError on cp1252 consoles. Idempotent; safe to run
# when this module is imported as a library too.
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except AttributeError:
        import codecs
        sys.stdout = codecs.getwriter("utf-8")(sys.stdout.buffer, "strict")
        sys.stderr = codecs.getwriter("utf-8")(sys.stderr.buffer, "strict")

SKILL_DIR = Path(__file__).resolve().parents[2]
REGISTRY_PATH = SKILL_DIR / "references" / "stahly_canonical_paths.md"
LOG_PATH = SKILL_DIR / "references" / "canonical_resolution_log.md"


# ============================================================
# Exceptions
# ============================================================

class CanonicalResourceNotFound(Exception):
    """Raised when a canonical resource cannot be resolved.

    Carries enough structured detail that the orchestrator can show the user
    exactly what was tried and ask where to look.
    """

    def __init__(
        self,
        resource_id: str,
        tried_hints: list[str],
        candidates_seen: list[dict],
        message: str = "",
    ):
        self.resource_id = resource_id
        self.tried_hints = tried_hints
        self.candidates_seen = candidates_seen
        if not message:
            message = self._format()
        super().__init__(message)

    def _format(self) -> str:
        lines = [f"Canonical resource '{self.resource_id}' not found."]
        lines.append("I looked in:")
        for h in self.tried_hints:
            lines.append(f"  - {h}")
        if self.candidates_seen:
            lines.append("")
            lines.append("Saw these candidates but they failed verification:")
            for c in self.candidates_seen:
                failed = ", ".join(c.get("failed_markers", []))
                lines.append(f"  - {c['path']}  (failed: {failed})")
        lines.append("")
        lines.append(
            "Where should I look? (Give me a path or a parent folder to "
            "search under. I'll re-resolve and update the registry.)"
        )
        return "\n".join(lines)


class RegistryParseError(Exception):
    """Raised when the YAML in stahly_canonical_paths.md can't be parsed."""


# ============================================================
# Registry parser (lightweight — no PyYAML dependency)
# ============================================================

# Each YAML block is fenced by ```yaml ... ``` and contains a flat or nested
# structure. We parse a minimal subset by hand to avoid pulling in PyYAML
# (which the skill otherwise doesn't need).

def _parse_registry(text: str) -> dict[str, dict]:
    """Return ``{resource_id: block}``. Each block is a dict mirroring the YAML."""
    blocks = re.findall(r"```yaml\n(.*?)\n```", text, flags=re.DOTALL)
    out: dict[str, dict] = {}
    for raw in blocks:
        try:
            block = _parse_yaml_block(raw)
        except Exception as e:
            raise RegistryParseError(f"Parse error in registry block:\n{raw[:200]}\n  -> {e}")
        rid = block.get("resource_id")
        if rid:
            out[rid] = block
    return out


def _parse_yaml_block(text: str) -> dict:
    """Parse a single YAML-ish block. Handles the limited subset used by the
    registry: scalars, lists of scalars, lists of dicts (with single-line
    key: value), and pipe-style multiline strings.
    """
    lines = text.splitlines()
    return _parse_lines(lines, indent=0)[0]


def _parse_lines(lines: list[str], indent: int) -> tuple[dict, int]:
    """Parse a block at the given indent level. Returns (dict, consumed)."""
    result: dict = {}
    i = 0
    while i < len(lines):
        line = lines[i]
        stripped = line.lstrip(" ")
        # Compute current indent
        cur_indent = len(line) - len(stripped)
        if not stripped or stripped.startswith("#"):
            i += 1
            continue
        if cur_indent < indent:
            break
        if cur_indent > indent:
            # shouldn't happen at top of a sub-block; advance
            i += 1
            continue

        # key: value
        m = re.match(r"^([A-Za-z_][A-Za-z0-9_]*)\s*:\s*(.*)$", stripped)
        if not m:
            # Could be a list item under a parent key — handled below
            i += 1
            continue
        key, val = m.group(1), m.group(2)

        if val == "":
            # Could be: nested dict, list (-), or multi-line
            # Look ahead
            j = i + 1
            while j < len(lines) and not lines[j].strip():
                j += 1
            if j >= len(lines):
                result[key] = None
                i = j
                continue
            next_line = lines[j]
            next_stripped = next_line.lstrip(" ")
            next_indent = len(next_line) - len(next_stripped)
            if next_indent <= indent:
                result[key] = None
                i = j
                continue
            if next_stripped.startswith("- "):
                # List
                items, consumed = _parse_list(lines[j:], indent=next_indent)
                result[key] = items
                i = j + consumed
            else:
                # Nested dict
                child, consumed = _parse_lines(lines[j:], indent=next_indent)
                result[key] = child
                i = j + consumed
        elif val == "|":
            # Multi-line literal block
            j = i + 1
            buf = []
            block_indent: int | None = None
            while j < len(lines):
                ln = lines[j]
                if not ln.strip():
                    buf.append("")
                    j += 1
                    continue
                ln_stripped = ln.lstrip(" ")
                ln_indent = len(ln) - len(ln_stripped)
                if block_indent is None:
                    if ln_indent <= indent:
                        break
                    block_indent = ln_indent
                if ln_indent < block_indent:
                    break
                buf.append(ln[block_indent:])
                j += 1
            # Trim trailing empties
            while buf and buf[-1] == "":
                buf.pop()
            result[key] = "\n".join(buf)
            i = j
        else:
            result[key] = _coerce_scalar(val)
            i += 1

    return result, i


def _parse_list(lines: list[str], indent: int) -> tuple[list, int]:
    """Parse a list at the given indent level. Returns (list, consumed)."""
    items: list = []
    i = 0
    while i < len(lines):
        line = lines[i]
        stripped = line.lstrip(" ")
        cur_indent = len(line) - len(stripped)
        if not stripped:
            i += 1
            continue
        if cur_indent < indent or not stripped.startswith("- "):
            break
        # Single-line list item: "- foo" or "- key: value"
        item_body = stripped[2:].lstrip(" ")
        if ":" in item_body and not item_body.startswith('"'):
            # Could be a dict item (single-line) — start a dict
            first_kv = re.match(r"^([A-Za-z_][A-Za-z0-9_]*)\s*:\s*(.*)$", item_body)
            if first_kv:
                key = first_kv.group(1)
                val = first_kv.group(2)
                item: dict = {key: _coerce_scalar(val) if val else None}
                # Continue reading sibling keys at deeper indent
                j = i + 1
                child_indent = indent + 2
                # Need to scan for keys at indent+2 (inside this list item)
                while j < len(lines):
                    sub = lines[j]
                    sub_stripped = sub.lstrip(" ")
                    sub_indent = len(sub) - len(sub_stripped)
                    if not sub_stripped:
                        j += 1
                        continue
                    if sub_indent < child_indent or sub_stripped.startswith("- "):
                        break
                    if sub_indent == child_indent:
                        m = re.match(r"^([A-Za-z_][A-Za-z0-9_]*)\s*:\s*(.*)$", sub_stripped)
                        if m:
                            k = m.group(1)
                            v = m.group(2)
                            if v == "":
                                # nested — recurse
                                child, consumed = _parse_lines(lines[j+1:], indent=child_indent + 2)
                                item[k] = child
                                j = j + 1 + consumed
                                continue
                            item[k] = _coerce_scalar(v)
                    j += 1
                items.append(item)
                i = j
                continue
        # Plain scalar list item
        items.append(_coerce_scalar(item_body))
        i += 1
    return items, i


_SCALAR_FALSE = {"false", "no", "off"}
_SCALAR_TRUE = {"true", "yes", "on"}
_SCALAR_NULL = {"null", "~", ""}


def _coerce_scalar(val: str) -> Any:
    val = val.strip()
    if not val:
        return None
    # Double-quoted: handle YAML escape sequences
    if val.startswith('"') and val.endswith('"') and len(val) >= 2:
        body = val[1:-1]
        return (
            body
            .replace('\\\\', '\x00')   # placeholder to avoid double-processing
            .replace('\\"', '"')
            .replace('\\n', '\n')
            .replace('\\t', '\t')
            .replace('\x00', '\\')
        )
    # Single-quoted: no escapes except ''
    if val.startswith("'") and val.endswith("'") and len(val) >= 2:
        return val[1:-1].replace("''", "'")
    # Bool/null
    lower = val.lower()
    if lower in _SCALAR_FALSE:
        return False
    if lower in _SCALAR_TRUE:
        return True
    if lower in _SCALAR_NULL:
        return None
    # Number
    try:
        if "." in val:
            return float(val)
        return int(val)
    except ValueError:
        pass
    return val


# ============================================================
# Fingerprint verification
# ============================================================

def _verify_markers(path: Path, markers: list[dict]) -> tuple[bool, list[str]]:
    """Check the candidate against every marker. Returns (ok, failed_marker_types)."""
    failed: list[str] = []
    for m in markers:
        t = m.get("type", "")
        try:
            if t == "filename_exact":
                if path.name != m.get("value", ""):
                    failed.append(t)
            elif t == "filename_pattern":
                if not fnmatch.fnmatch(path.name, m.get("value", "")):
                    failed.append(t)
            elif t == "xlsx_sheet_name":
                if not _xlsx_has_sheet(path, m.get("value", "")):
                    failed.append(t)
            elif t == "xlsx_cell_value":
                if not _xlsx_cell_contains(
                    path, m.get("sheet", ""), m.get("cell", ""), m.get("contains", "")
                ):
                    failed.append(t)
            elif t == "docx_paragraph_starts_with":
                if not _docx_paragraph_starts_with(path, m.get("value", "")):
                    failed.append(t)
            elif t == "pdf_text_contains":
                if not _pdf_text_contains(path, m.get("value", "")):
                    failed.append(t)
            elif t == "is_directory":
                if not path.is_dir():
                    failed.append(t)
            elif t == "directory_contains_pattern":
                if not _dir_has_match(path, m.get("value", "")):
                    failed.append(t)
            elif t == "image_dimensions":
                if not _image_matches(path, m.get("width"), m.get("height")):
                    failed.append(t)
            else:
                failed.append(f"unknown_marker:{t}")
        except Exception:
            failed.append(t)
    return (not failed, failed)


def _xlsx_has_sheet(path: Path, sheet: str) -> bool:
    from openpyxl import load_workbook
    try:
        wb = load_workbook(path, read_only=True, data_only=False)
        return sheet in wb.sheetnames
    except Exception:
        return False


def _xlsx_cell_contains(path: Path, sheet: str, cell: str, needle: str) -> bool:
    from openpyxl import load_workbook
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            return False
        v = wb[sheet][cell].value
        if v is None:
            return False
        return needle.lower() in str(v).lower()
    except Exception:
        return False


def _docx_paragraph_starts_with(path: Path, prefix: str) -> bool:
    try:
        from docx import Document
        doc = Document(str(path))
        return any(p.text.strip().startswith(prefix) for p in doc.paragraphs)
    except Exception:
        return False


def _pdf_text_contains(path: Path, needle: str) -> bool:
    """Lightweight PDF text check. Tries pypdf, then pdfplumber, then bytes fallback."""
    try:
        from pypdf import PdfReader  # type: ignore
        try:
            reader = PdfReader(str(path))
            text = ""
            for page in reader.pages[:3]:  # First 3 pages is enough for a fingerprint
                text += page.extract_text() or ""
            return needle.lower() in text.lower()
        except Exception:
            pass
    except ImportError:
        pass
    try:
        with open(path, "rb") as f:
            raw = f.read(200_000)
        return needle.lower().encode("utf-8") in raw.lower()
    except Exception:
        return False


def _dir_has_match(path: Path, pattern: str) -> bool:
    try:
        return any(fnmatch.fnmatch(p.name, pattern) for p in path.iterdir())
    except Exception:
        return False


def _image_matches(path: Path, w: int | None, h: int | None) -> bool:
    if w is None or h is None:
        return True
    try:
        from PIL import Image
        with Image.open(path) as im:
            return im.size == (w, h)
    except Exception:
        return False


# ============================================================
# Resolver
# ============================================================

@dataclass
class ResolutionTrace:
    resource_id: str
    cached_path_tried: str | None = None
    cached_path_failed_markers: list[str] = field(default_factory=list)
    hints_tried: list[str] = field(default_factory=list)
    candidates_seen: list[dict] = field(default_factory=list)
    resolved_path: str | None = None
    resolved_at_new_location: bool = False


def resolve(resource_id: str, *, force_refresh: bool = False) -> Path:
    """Resolve a canonical resource. Returns its current path on disk.

    Resolution order:
        1. The cached ``resolved_path``, if it exists and fingerprint matches.
        2. Each entry in ``search_hints``, in order. First fingerprint match
           wins; the registry is updated with the new path.
        3. ``CanonicalResourceNotFound`` raised. No silent fallback.

    Set ``force_refresh=True`` to skip step 1 and re-search.
    """
    registry = _load_registry()
    if resource_id not in registry:
        raise CanonicalResourceNotFound(
            resource_id,
            tried_hints=[],
            candidates_seen=[],
            message=(
                f"Unknown canonical resource_id '{resource_id}'. "
                f"Known ids: {sorted(registry.keys())}"
            ),
        )

    block = registry[resource_id]
    markers = block.get("content_markers", []) or []
    hints = block.get("search_hints", []) or []
    cached = block.get("resolved_path")

    trace = ResolutionTrace(resource_id=resource_id)

    # Step 1: try cached path
    if cached and not force_refresh:
        cached_path = Path(cached)
        trace.cached_path_tried = str(cached_path)
        if cached_path.exists():
            ok, failed = _verify_markers(cached_path, markers)
            if ok:
                trace.resolved_path = str(cached_path)
                return cached_path
            trace.cached_path_failed_markers = failed
            trace.candidates_seen.append({"path": str(cached_path), "failed_markers": failed})
        else:
            trace.cached_path_failed_markers = ["path_does_not_exist"]

    # Step 2: walk search hints
    for hint in hints:
        trace.hints_tried.append(hint)
        for candidate in _glob_hint(hint):
            ok, failed = _verify_markers(candidate, markers)
            if ok:
                # Found it. Update registry and log.
                update_resolved_path(resource_id, candidate)
                trace.resolved_path = str(candidate)
                trace.resolved_at_new_location = (cached != str(candidate))
                if trace.resolved_at_new_location:
                    _log_resolution(resource_id, cached, candidate)
                return candidate
            trace.candidates_seen.append({"path": str(candidate), "failed_markers": failed})

    # Step 3: hard fail
    raise CanonicalResourceNotFound(
        resource_id=resource_id,
        tried_hints=trace.hints_tried,
        candidates_seen=trace.candidates_seen,
    )


def resolve_all() -> dict[str, Path]:
    """Resolve every resource in the registry. Returns {resource_id: Path}.
    Raises CanonicalResourceNotFound on the first miss (caller decides whether
    to continue with remaining resources or escalate immediately).
    """
    registry = _load_registry()
    out: dict[str, Path] = {}
    for rid in registry:
        out[rid] = resolve(rid)
    return out


def _glob_hint(hint: str) -> list[Path]:
    """Expand a glob hint into a list of existing paths.

    Supports ``**`` recursion. UNC paths are passed through to ``glob.glob``
    with ``recursive=True``. We sort results for determinism.
    """
    try:
        results = glob.glob(hint, recursive=True)
    except Exception:
        return []
    paths = [Path(r) for r in results if Path(r).exists()]
    paths.sort()
    return paths


# ============================================================
# Registry I/O
# ============================================================

_registry_cache: dict[str, dict] | None = None


def _load_registry(force: bool = False) -> dict[str, dict]:
    global _registry_cache
    if _registry_cache is not None and not force:
        return _registry_cache
    text = REGISTRY_PATH.read_text(encoding="utf-8")
    _registry_cache = _parse_registry(text)
    return _registry_cache


def update_resolved_path(resource_id: str, new_path: Path | str) -> None:
    """Rewrite the resolved_path inside the named YAML block of the registry.

    Preserves everything else in the file. Updates resolved_at to today.
    Invalidates the registry cache.
    """
    global _registry_cache
    new_str = str(new_path)
    today = _dt.date.today().isoformat()
    text = REGISTRY_PATH.read_text(encoding="utf-8")

    # Find the YAML block containing resource_id: <id>
    pattern = re.compile(
        r"(```yaml\n(?:(?!```).)*?resource_id:\s*" + re.escape(resource_id)
        + r"(?:(?!```).)*?\n```)",
        flags=re.DOTALL,
    )
    m = pattern.search(text)
    if not m:
        raise RegistryParseError(
            f"Could not find YAML block for resource_id '{resource_id}' in registry"
        )

    block = m.group(1)
    # Replace resolved_path line
    new_block = re.sub(
        r'^resolved_path:.*$',
        f'resolved_path: "{_escape_yaml_string(new_str)}"',
        block,
        count=1,
        flags=re.MULTILINE,
    )
    # If resolved_path didn't exist, insert it before the closing ```
    if new_block == block:
        new_block = block.replace(
            "```",
            f'resolved_path: "{_escape_yaml_string(new_str)}"\n```',
            1,
        )

    # Replace or insert resolved_at
    if re.search(r"^resolved_at:", new_block, flags=re.MULTILINE):
        new_block = re.sub(
            r'^resolved_at:.*$',
            f'resolved_at: "{today}"',
            new_block,
            count=1,
            flags=re.MULTILINE,
        )
    else:
        # Insert after resolved_path
        new_block = re.sub(
            r'(resolved_path:.*\n)',
            r'\1' + f'resolved_at: "{today}"\n',
            new_block,
            count=1,
        )

    text = text[:m.start()] + new_block + text[m.end():]
    REGISTRY_PATH.write_text(text, encoding="utf-8")
    _registry_cache = None


def _escape_yaml_string(s: str) -> str:
    # We use double-quoted YAML strings; backslashes need escaping
    return s.replace("\\", "\\\\").replace('"', '\\"')


def _log_resolution(resource_id: str, old_path: str | None, new_path: Path) -> None:
    """Append a one-line entry to the resolution log."""
    today = _dt.date.today().isoformat()
    if not LOG_PATH.exists():
        LOG_PATH.write_text(
            "# Canonical Resolution Log\n\n"
            "Auto-appended whenever the resolver finds a resource at a new "
            "location. Each line records when a path drifted.\n\n",
            encoding="utf-8",
        )
    with LOG_PATH.open("a", encoding="utf-8") as f:
        f.write(f"- {today} `{resource_id}`: `{old_path or '(none)'}` → `{new_path}`\n")


# ============================================================
# CLI for ad-hoc resolution + diagnostic
# ============================================================

def _cli(argv: list[str]) -> int:
    if len(argv) < 2 or argv[1] in ("-h", "--help"):
        print(
            "Usage:\n"
            "  python -m scripts.lib.canonical <resource_id>          # resolve one\n"
            "  python -m scripts.lib.canonical --all                  # resolve all\n"
            "  python -m scripts.lib.canonical <resource_id> --refresh # skip cached_path",
            file=sys.stderr,
        )
        return 2
    if argv[1] == "--all":
        try:
            out = resolve_all()
            for rid, p in out.items():
                print(f"{rid}\t{p}")
            return 0
        except CanonicalResourceNotFound as e:
            print(str(e), file=sys.stderr)
            return 1
    rid = argv[1]
    force = "--refresh" in argv[2:]
    try:
        p = resolve(rid, force_refresh=force)
        print(p)
        return 0
    except CanonicalResourceNotFound as e:
        print(str(e), file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(_cli(sys.argv))
