"""
Stahly bid workbook — shared library module.

All `bid_*.py` CLI scripts import from this module. The contract (sheet
fingerprint, row layout, column layout, rate fallback table, ROUNDUP rule,
hard rules) is documented in `references/bid_workbook_spec.md`.

Two execution paths
-------------------

**Reads** (`read_workbook`, `compute_total`, `fee_table_payload`,
`detect_bid_sheet`) use openpyxl read-only. Safe — openpyxl reads x14 data
validations even though it can't write them back.

**Writes** (`apply_change`, `unhide_all`, plus the COM-prefixed helpers) use
Microsoft Excel via late-binding `win32com.client.Dispatch` because **openpyxl
strips x14-namespace data validations on save**, which silently kills every
drop-down in the workbook. Writes always operate on a local working copy in
``%TEMP%`` and push back to the original location on success.

Public API
----------
    detect_bid_sheet(wb)                        -> openpyxl worksheet
    read_workbook(path)                         -> audit dict
    compute_total(path)                         -> totals dict
    fee_table_payload(path, title, lead)        -> build.py payload
    apply_change(path, row, col, value)         -> result dict (COM write)
    unhide_all(path)                            -> counts (COM write)
    diff_totals(path, snapshot)                 -> deltas + alerts

Hard rules enforced here:
  - Resolve paths via canonical.resolve when called without an explicit path.
  - Use Excel COM (late-binding) for all writes.
  - Work on local copies under %TEMP%; push back on success.
  - Never overwrite formula cells.
  - Never insert or delete rows/columns.
  - Use cell.Value = None for clears (works on merged cells).
  - Use datetime.datetime for date values (COM rejects datetime.date).
  - Refuse to operate on workbooks that fail the fingerprint check.
  - Surface every failure to the user — no silent fallback.
"""
from __future__ import annotations

import datetime as _dt
import math
import os
import shutil
import sys
import tempfile
from contextlib import contextmanager
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string

# ============================================================
# CONSTANTS (locked — see references/bid_workbook_spec.md)
# ============================================================

# Phase layout: (header_row, first_task_row, last_task_row, subtotal_row)
# 2026 template: 12 phases, each 12 rows.
PHASES: list[tuple[int, int, int, int]] = [
    (8,   9,   18,  19),   # Phase 1
    (20,  21,  30,  31),   # Phase 2
    (32,  33,  42,  43),   # Phase 3
    (44,  45,  54,  55),   # Phase 4
    (56,  57,  66,  67),   # Phase 5
    (68,  69,  78,  79),   # Phase 6
    (80,  81,  90,  91),   # Phase 7
    (92,  93,  102, 103),  # Phase 8
    (104, 105, 114, 115),  # Phase 9
    (116, 117, 126, 127),  # Phase 10
    (128, 129, 138, 139),  # Phase 11
    (140, 141, 150, 151),  # Phase 12
]

# Staff hour columns — 2026 template uses G through Q (11 slots; older was G–K).
STAFF_COLS = ["G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q"]

# Expense columns
COL_MILEAGE = "S"
COL_MILEAGE_RATE = "T"       # row 7 holds the rate
COL_PER_DIEM_DAYS = "U"
COL_PER_DIEM_RATE = "V"      # row 7 holds the rate
COL_UAV_FEE = "W"            # also GPS/scanner/etc. equipment day rates
COL_CONSULTANT = "X"         # auto +5%
COL_OTHER = "Y"

COL_TASK_DESC = "B"
COL_TASK_TOTAL = "F"
COL_LABOR_SUBTOTAL = "R"
COL_EXPENSE_SUBTOTAL = "Z"

# Header cells (project metadata)
CELL_CLIENT = "F1"
CELL_PROJECT = "F2"
CELL_PREPARED_BY = "C3"
CELL_CHECKED_BY = "E3"
CELL_DATE = "C4"
CELL_RATE_SHEET_SELECTOR = "C6"

# Grand-total rows
ROW_GRAND_TOTAL = 152
ROW_ROUNDED_TOTAL = 153

# Pre-existing template defects (Employee List rows 74-81 VLOOKUPs).
# Used by verify_assets.py to suppress benign #N/A noise.
KNOWN_PRISTINE_ERROR_RANGES = [
    {"sheet": "Employee List", "rows": (74, 81), "cols": "C:J"},
]

# Hardcoded fallback rates (2026 Standard) — used only when Rate Sheet is
# missing/unreadable. Update yearly.
FALLBACK_RATES = {
    "LPS1": 156.00, "LPS2": 167.00, "LPS3": 178.00, "LPS4": 188.00,
    "LPS5": 198.00, "LPS6": 208.00, "LPS7": 212.00,
    "LST1": 120.00, "LST2": 126.00, "LST3": 133.00, "LST4": 142.00,
    "LST5": 149.00, "LST6": 156.00,
    "LSI1": 126.00, "LSI2": 133.00, "LSI3": 142.00, "LSI4": 149.00,
    "EPE1": 163.00, "EPE2": 171.00, "EPE3": 186.00, "EPE4": 202.00,
    "EPE5": 216.00, "EPE6": 224.00, "EPE7": 230.00, "EPE8": 246.00,
}

# Hardcoded fallback expense constants (2026)
FALLBACK_MILEAGE_RATE = 0.75
FALLBACK_PER_DIEM_RATE = 58.00


# ============================================================
# Helpers
# ============================================================

def _is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _numeric(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    return 0.0


def _roundup_1k(amount: float) -> int:
    """Match Excel ROUNDUP(F152, -3) — round UP to nearest $1,000."""
    return int(math.ceil(amount / 1000.0)) * 1000


def _resolve_path(path: Optional[str | Path]) -> Path:
    """If path is None, resolve via canonical. Otherwise validate and return."""
    if path is None:
        resolve = _import_canonical_resolve()
        return resolve("bid_template")
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(
            f"Bid workbook not found: {p}\n"
            "If the file moved, update references/stahly_canonical_paths.md or "
            "pass the new path explicitly."
        )
    return p


def _import_canonical_resolve():
    """Import canonical.resolve. The canonical module lives at scripts/lib/
    and the bid_workbook scripts can be invoked from several entry points;
    try the common ones."""
    import sys as _sys
    _scripts_dir = Path(__file__).resolve().parent
    if str(_scripts_dir) not in _sys.path:
        _sys.path.insert(0, str(_scripts_dir))
    try:
        from lib.canonical import resolve  # type: ignore
        return resolve
    except ImportError:
        pass
    # Fallback: try the package path from skill root
    _skill_dir = _scripts_dir.parent
    if str(_skill_dir) not in _sys.path:
        _sys.path.insert(0, str(_skill_dir))
    from scripts.lib.canonical import resolve  # type: ignore
    return resolve


# ============================================================
# Sheet fingerprint detection
# ============================================================

def detect_bid_sheet(wb):
    """Return the first worksheet that matches the Stahly 2026 bid workbook
    fingerprint. Raises ValueError if none match.

    Fingerprint (all must be true on a single sheet). All checks are
    *structural* — they pass on both blank and filled workbooks:
      1. B7 starts with "Project Tasks" (case-insensitive)
      2. B152 starts with "Total Project Costs"
      3. B153 starts with "TOTAL PROJECT FEE"
      4. A7 contains "Phase - Blue" (the label cell present in every template)
      5. Row 7 cols G-Q have at least one formula referencing 'Rate Sheet'
    """
    candidates = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        a7 = ws["A7"].value
        b7 = ws["B7"].value
        b_total = ws[f"B{ROW_GRAND_TOTAL}"].value
        b_round = ws[f"B{ROW_ROUNDED_TOTAL}"].value

        if not isinstance(b7, str) or not b7.strip().lower().startswith("project tasks"):
            continue
        if not isinstance(b_total, str) or not b_total.strip().lower().startswith("total project costs"):
            continue
        if not isinstance(b_round, str) or not b_round.strip().lower().startswith("total project fee"):
            continue
        if not isinstance(a7, str) or "phase - blue" not in a7.lower():
            continue

        row7_formulas = [ws[f"{c}7"].value for c in STAFF_COLS]
        if not any(_is_formula(f) and "Rate Sheet" in f for f in row7_formulas):
            continue

        candidates.append(ws)

    if not candidates:
        raise ValueError(
            "No sheet in this workbook matches the Stahly 2026 bid workbook fingerprint.\n"
            "Required (all on a single sheet):\n"
            "  - A7 contains 'Phase - Blue'\n"
            "  - B7 starts with 'Project Tasks'\n"
            "  - B152 starts with 'Total Project Costs'\n"
            "  - B153 starts with 'TOTAL PROJECT FEE'\n"
            "  - Row 3 cols G-Q has at least one staff name\n"
            "  - Row 7 cols G-Q has a formula referencing 'Rate Sheet'\n"
            "See references/bid_workbook_spec.md."
        )
    return candidates[0]


# ============================================================
# Rate sheet read
# ============================================================

def _resolve_rates(ws_formulas, ws_cached) -> tuple[dict, dict]:
    """Return ({staff_col: hourly_rate}, {col: source_label}).

    Resolution order per staff column:
      1. row-7 formula-sheet value if numeric (constant override)
      2. row-7 cached value if numeric (formula resolved by Excel)
      3. row-6 cached code → FALLBACK_RATES
      4. zero (silent underbid trap — caller must warn)
    """
    rates: dict = {}
    sources: dict = {}

    for col in STAFF_COLS:
        raw_f = ws_formulas[f"{col}7"].value
        raw_v = ws_cached[f"{col}7"].value

        if isinstance(raw_f, (int, float)) and raw_f > 0:
            rates[col] = float(raw_f)
            sources[col] = "workbook"
        elif isinstance(raw_v, (int, float)) and raw_v > 0:
            rates[col] = float(raw_v)
            sources[col] = "workbook"
        else:
            code = ws_cached[f"{col}6"].value
            if isinstance(code, str) and code.strip().upper() in FALLBACK_RATES:
                rates[col] = FALLBACK_RATES[code.strip().upper()]
                sources[col] = "fallback"
            else:
                rates[col] = 0.0
                sources[col] = "unresolved"

    def _pick(col_letter, default):
        raw_f = ws_formulas[f"{col_letter}7"].value
        raw_v = ws_cached[f"{col_letter}7"].value
        if isinstance(raw_f, (int, float)) and raw_f > 0:
            return float(raw_f), "workbook"
        if isinstance(raw_v, (int, float)) and raw_v > 0:
            return float(raw_v), "workbook"
        return float(default), "fallback"

    mileage_rate, mileage_src = _pick(COL_MILEAGE_RATE, FALLBACK_MILEAGE_RATE)
    per_diem_rate, per_diem_src = _pick(COL_PER_DIEM_RATE, FALLBACK_PER_DIEM_RATE)

    rates["_mileage"] = mileage_rate
    rates["_per_diem"] = per_diem_rate
    sources["_mileage"] = mileage_src
    sources["_per_diem"] = per_diem_src

    return rates, sources


# ============================================================
# Silent-underbid warnings
# ============================================================

def _hours_in_column(ws, col: str) -> float:
    total = 0.0
    for _, first, last, _ in PHASES:
        for row in range(first, last + 1):
            total += _numeric(ws[f"{col}{row}"].value)
    return total


def _underbid_warnings(ws, rates: dict, sources: dict) -> list[str]:
    warnings: list[str] = []
    for col in STAFF_COLS:
        staff = ws[f"{col}3"].value
        code = ws[f"{col}6"].value  # in case of labor-code override
        rate = rates.get(col, 0.0)
        hours = _hours_in_column(ws, col)
        src = sources.get(col, "unresolved")

        # Override-pattern is legitimate when a labor code is present in row 6
        # even if row 3 is blank — don't false-warn on that case.
        if not staff and hours > 0:
            if isinstance(code, str) and code.strip().upper() in FALLBACK_RATES:
                continue  # labor-code override in play, rate resolved
            warnings.append(
                f"Column {col} has {hours:g} hours but row-3 staff name is blank "
                f"AND row-6 has no recognized labor code — those hours will resolve "
                f"to $0 (silent underbid). Set {col}3 or {col}6."
            )
            continue

        if staff and rate == 0 and hours > 0:
            warnings.append(
                f"Column {col} ({staff}) has {hours:g} hours but rate is $0. "
                f"Row-7 lookup failed (source: {src}). Check {col}6 labor code "
                f"and Rate Sheet, or hand-enter the rate at {col}7."
            )
            continue

        if staff and src == "fallback":
            warnings.append(
                f"Column {col} ({staff}) rate ${rate:.0f}/hr from FALLBACK_RATES "
                f"table — verify against live Rate Sheet for the current year."
            )
    return warnings


# ============================================================
# Workbook read (audit)
# ============================================================

def read_workbook(path: str | Path | None = None) -> dict:
    """Audit-level read. Returns phases, tasks, rates, expense detail,
    grand totals, and warnings."""
    p = _resolve_path(path)
    wb = load_workbook(p, data_only=False)
    wb_v = load_workbook(p, data_only=True)

    ws = detect_bid_sheet(wb)
    ws_v = wb_v[ws.title]

    rates, sources = _resolve_rates(ws, ws_v)
    warnings = _underbid_warnings(ws, rates, sources)

    staff_srcs = {sources[c] for c in STAFF_COLS}
    if staff_srcs == {"workbook"}:
        rate_source = "workbook"
    elif "unresolved" in staff_srcs:
        rate_source = "unresolved" if len(staff_srcs) == 1 else "mixed"
    elif "fallback" in staff_srcs:
        rate_source = "fallback" if len(staff_srcs) == 1 else "mixed"
    else:
        rate_source = "mixed"

    metadata = {
        "workbook_path": str(p),
        "sheet_name": ws.title,
        "client": ws[CELL_CLIENT].value,
        "project": ws[CELL_PROJECT].value,
        "prepared_by": ws[CELL_PREPARED_BY].value,
        "checked_by": ws[CELL_CHECKED_BY].value,
        "date": str(ws[CELL_DATE].value) if ws[CELL_DATE].value else None,
        "rate_sheet_selector": ws[CELL_RATE_SHEET_SELECTOR].value,
        "rate_source": rate_source,
        "rate_sources": {c: sources[c] for c in STAFF_COLS},
        "staff": {c: ws[f"{c}3"].value for c in STAFF_COLS},
        "labor_codes": {c: ws[f"{c}6"].value for c in STAFF_COLS},
        "staff_rates": {c: rates[c] for c in STAFF_COLS},
        "mileage_rate": rates["_mileage"],
        "mileage_rate_source": sources["_mileage"],
        "per_diem_rate": rates["_per_diem"],
        "per_diem_rate_source": sources["_per_diem"],
    }

    phases_out = []
    total_labor = 0.0
    total_expense = 0.0

    for idx, (hdr, first, last, sub) in enumerate(PHASES, 1):
        phase_name_raw = ws[f"B{hdr}"].value or f"Phase {idx}"
        tasks = []
        phase_labor = 0.0
        phase_expense = 0.0

        for row in range(first, last + 1):
            desc = ws[f"B{row}"].value
            if not desc or _is_formula(desc):
                # Still tally any hours/expenses on rows without a description —
                # the user may have left B blank but populated hours.
                row_hours = {c: _numeric(ws[f"{c}{row}"].value) for c in STAFF_COLS}
                row_labor_total = sum(row_hours[c] * rates[c] for c in STAFF_COLS)
                if row_labor_total == 0 and all(
                    _numeric(ws[f"{c}{row}"].value) == 0
                    for c in (COL_MILEAGE, COL_PER_DIEM_DAYS, COL_UAV_FEE,
                              COL_CONSULTANT, COL_OTHER)
                ):
                    continue
                # Otherwise fall through and account for the row
                desc = f"(row {row})"

            hours = {c: _numeric(ws[f"{c}{row}"].value) for c in STAFF_COLS}
            labor = sum(hours[c] * rates[c] for c in STAFF_COLS)

            miles = _numeric(ws[f"{COL_MILEAGE}{row}"].value)
            per_diem_days = _numeric(ws[f"{COL_PER_DIEM_DAYS}{row}"].value)
            uav_fee = _numeric(ws[f"{COL_UAV_FEE}{row}"].value)
            consultant = _numeric(ws[f"{COL_CONSULTANT}{row}"].value)
            other = _numeric(ws[f"{COL_OTHER}{row}"].value)

            expense = (
                miles * rates["_mileage"]
                + per_diem_days * rates["_per_diem"]
                + uav_fee
                + consultant * 1.05
                + other
            )

            tasks.append({
                "row": row,
                "description": desc,
                "hours": {c: hours[c] for c in STAFF_COLS if hours[c] > 0},
                "labor": round(labor, 2),
                "expense_detail": {
                    "mileage_miles": miles,
                    "mileage_cost": round(miles * rates["_mileage"], 2),
                    "per_diem_days": per_diem_days,
                    "per_diem_cost": round(per_diem_days * rates["_per_diem"], 2),
                    "uav_fee": uav_fee,
                    "consultant": consultant,
                    "consultant_with_markup": round(consultant * 1.05, 2),
                    "other": other,
                },
                "expense_total": round(expense, 2),
                "task_total": round(labor + expense, 2),
            })

            phase_labor += labor
            phase_expense += expense

        phase_name = phase_name_raw.strip() if isinstance(phase_name_raw, str) else phase_name_raw
        phases_out.append({
            "index": idx,
            "name": phase_name,
            "header_row": hdr,
            "task_rows": [first, last],
            "subtotal_row": sub,
            "tasks": tasks,
            "labor_subtotal": round(phase_labor, 2),
            "expense_subtotal": round(phase_expense, 2),
            "phase_total": round(phase_labor + phase_expense, 2),
        })
        total_labor += phase_labor
        total_expense += phase_expense

    grand = total_labor + total_expense
    rounded = _roundup_1k(grand)

    return {
        "metadata": metadata,
        "phases": phases_out,
        "totals": {
            "labor": round(total_labor, 2),
            "expenses": round(total_expense, 2),
            "grand": round(grand, 2),
            "rounded_1k": rounded,
        },
        "warnings": warnings,
    }


def compute_total(path: str | Path | None = None) -> dict:
    full = read_workbook(path)
    return full["totals"]


# ============================================================
# Excel COM write context — used by all writers
# ============================================================

@contextmanager
def _excel_workbook(path: Path, read_only: bool = False):
    """Open the workbook via late-binding Excel COM, yield (wb, app).

    Uses late-binding `Dispatch` (not `gencache.EnsureDispatch`) because
    Stahly's Office typelib breaks the cache generator on this machine.
    Always sets Visible=False, DisplayAlerts=False. Caller is responsible
    for saving — the context manager closes without saving.
    """
    try:
        import win32com.client as win32  # type: ignore
        import pythoncom  # type: ignore
    except ImportError as e:
        raise RuntimeError(
            "Excel COM writes require pywin32. Install with: pip install pywin32"
        ) from e

    pythoncom.CoInitialize()
    app = win32.Dispatch("Excel.Application")
    app.Visible = False
    app.DisplayAlerts = False
    try:
        wb = app.Workbooks.Open(str(path), ReadOnly=read_only)
        try:
            yield wb, app
        finally:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
    finally:
        try:
            app.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()


def _detect_sheet_via_com(wb) -> object:
    """COM-based version of detect_bid_sheet. Returns the Sheet object."""
    for ws_com in wb.Worksheets:
        try:
            a7 = ws_com.Range("A7").Value
            b7 = ws_com.Range("B7").Value
            b_total = ws_com.Range(f"B{ROW_GRAND_TOTAL}").Value
            b_round = ws_com.Range(f"B{ROW_ROUNDED_TOTAL}").Value
        except Exception:
            continue
        if not isinstance(a7, str) or "phase - blue" not in a7.lower():
            continue
        if not isinstance(b7, str) or not b7.strip().lower().startswith("project tasks"):
            continue
        if not isinstance(b_total, str) or not b_total.strip().lower().startswith("total project costs"):
            continue
        if not isinstance(b_round, str) or not b_round.strip().lower().startswith("total project fee"):
            continue
        # Row 7 must have at least one rate-formula present (structural check;
        # row 3 staff names are content, not structure)
        has_rate_formula = False
        for c in STAFF_COLS:
            f = ws_com.Range(f"{c}7").Formula
            if isinstance(f, str) and f.startswith("=") and "Rate Sheet" in f:
                has_rate_formula = True
                break
        if not has_rate_formula:
            continue
        return ws_com
    raise ValueError(
        "No sheet matches the Stahly 2026 bid workbook fingerprint via COM."
    )


@contextmanager
def _local_working_copy(network_path: Path):
    """Copy the network file to %TEMP%, yield the local path, push back on
    clean exit. On exception, the local copy is preserved for inspection but
    not pushed back."""
    tmp_dir = Path(tempfile.mkdtemp(prefix="stahly_bid_"))
    local = tmp_dir / network_path.name
    shutil.copy2(network_path, local)
    pushed_back = False
    try:
        yield local
        # Push back on success
        try:
            shutil.copy2(local, network_path)
            pushed_back = True
        except PermissionError:
            print(
                f"\nCANNOT PUSH BACK: {network_path.name} is locked at the destination.\n"
                f"Working copy preserved at: {local}\n"
                f"Close the file (probably open in Excel) and copy manually.",
                file=sys.stderr,
            )
            raise
    finally:
        if pushed_back:
            try:
                shutil.rmtree(tmp_dir)
            except Exception:
                pass


# ============================================================
# Safe cell write via Excel COM
# ============================================================

def apply_change(path: str | Path | None, row: int, col_letter: str, value) -> dict:
    """Modify a single cell via Excel COM. Refuses to overwrite formulas.
    Refuses to operate on subtotal or grand-total rows.

    `value` may be: str, int, float, bool, datetime.datetime, or None
    (None clears the cell — works on merged cells, unlike ClearContents).
    """
    p = _resolve_path(path)

    valid_task_rows = set()
    for _, first, last, _ in PHASES:
        for r in range(first, last + 1):
            valid_task_rows.add(r)
    if row not in valid_task_rows:
        raise ValueError(
            f"Row {row} is not a task row. Valid: 9-18, 21-30, 33-42, 45-54, "
            f"57-66, 69-78, 81-90, 93-102, 105-114, 117-126, 129-138, 141-150. "
            "Header, subtotal, and grand-total rows are not editable."
        )

    col = col_letter.upper()
    cell_ref = f"{col}{row}"

    # COM cannot accept datetime.date; convert
    if isinstance(value, _dt.date) and not isinstance(value, _dt.datetime):
        value = _dt.datetime(value.year, value.month, value.day)

    with _local_working_copy(p) as local:
        with _excel_workbook(local) as (wb, app):
            ws = _detect_sheet_via_com(wb)
            cell = ws.Range(cell_ref)
            before_formula = cell.Formula
            before_value = cell.Value
            if isinstance(before_formula, str) and before_formula.startswith("="):
                raise ValueError(
                    f"Refusing to overwrite formula in {cell_ref}: {before_formula!r}"
                )
            cell.Value = value
            wb.Save()
    return {
        "cell": cell_ref,
        "sheet": "Bid Sheet Template",
        "before": before_value,
        "after": value,
        "saved_to": str(p),
    }


def apply_changes(path: str | Path | None, changes: list[dict]) -> dict:
    """Batch version of apply_change. `changes` is a list of
    ``{"row": int, "col": str, "value": Any}`` dicts. Opens the workbook
    once, applies every change inside a single COM session, then saves.

    Refuses the whole batch if any single change targets a formula cell
    or a non-task row.
    """
    p = _resolve_path(path)

    valid_task_rows = set()
    for _, first, last, _ in PHASES:
        for r in range(first, last + 1):
            valid_task_rows.add(r)

    # Also allow header/metadata cells (row 1-7) for the bootstrap workflow
    # but only specific named cells.
    allowed_header_cells = {
        CELL_CLIENT, CELL_PROJECT,
        CELL_PREPARED_BY, CELL_CHECKED_BY, CELL_DATE,
        CELL_RATE_SHEET_SELECTOR,
    }
    # Plus row-3 staff name columns G-Q
    allowed_header_cells.update(f"{c}3" for c in STAFF_COLS)
    # Plus phase header B-column cells (rename phases)
    allowed_header_cells.update(f"B{hdr}" for hdr, *_ in PHASES)
    # Plus task-manager C-column for each phase header
    allowed_header_cells.update(f"C{hdr}" for hdr, *_ in PHASES)
    # Plus the row-7 mileage/per-diem rate inputs
    allowed_header_cells.update({"T7", "V7"})
    # Plus the row-6 labor-code override cells
    allowed_header_cells.update(f"{c}6" for c in STAFF_COLS)

    # Pre-flight validation
    for ch in changes:
        row = ch["row"]
        col = ch["col"].upper()
        cell_ref = f"{col}{row}"
        if row not in valid_task_rows and cell_ref not in allowed_header_cells:
            raise ValueError(
                f"Change targets {cell_ref}, which is not a task row nor an "
                f"allowed header/metadata cell."
            )

    with _local_working_copy(p) as local:
        with _excel_workbook(local) as (wb, app):
            ws = _detect_sheet_via_com(wb)

            # Pre-flight formula check (inside the COM session, against live cells)
            bad = []
            for ch in changes:
                cell_ref = f"{ch['col'].upper()}{ch['row']}"
                f = ws.Range(cell_ref).Formula
                if isinstance(f, str) and f.startswith("="):
                    bad.append((cell_ref, f))
            if bad:
                msg = "Refusing batch — these target cells contain formulas:\n"
                msg += "\n".join(f"  {c}: {f}" for c, f in bad)
                raise ValueError(msg)

            applied = []
            for ch in changes:
                cell_ref = f"{ch['col'].upper()}{ch['row']}"
                v = ch["value"]
                if isinstance(v, _dt.date) and not isinstance(v, _dt.datetime):
                    v = _dt.datetime(v.year, v.month, v.day)
                before = ws.Range(cell_ref).Value
                ws.Range(cell_ref).Value = v
                applied.append({"cell": cell_ref, "before": before, "after": v})

            wb.Save()
    return {"saved_to": str(p), "count": len(applied), "applied": applied}


# ============================================================
# Unhide all (via Excel COM so x14 validations survive)
# ============================================================

def unhide_all(path: str | Path | None = None) -> dict:
    """Unhide every row and column across every sheet. Uses Excel COM so
    data validations are preserved."""
    p = _resolve_path(path)
    by_sheet: dict[str, dict] = {}
    rows_unhidden = 0
    cols_unhidden = 0

    with _local_working_copy(p) as local:
        with _excel_workbook(local) as (wb, app):
            for ws_com in wb.Worksheets:
                sheet_rows = 0
                sheet_cols = 0
                try:
                    used = ws_com.UsedRange
                    rows = used.EntireRow
                    rows.Hidden = False
                    cols = used.EntireColumn
                    cols.Hidden = False
                except Exception:
                    pass
                # Iterate to count (COM doesn't expose a quick count)
                try:
                    last_row = ws_com.UsedRange.Row + ws_com.UsedRange.Rows.Count - 1
                    for r in range(1, last_row + 1):
                        if ws_com.Rows(r).Hidden:
                            ws_com.Rows(r).Hidden = False
                            sheet_rows += 1
                    last_col = ws_com.UsedRange.Column + ws_com.UsedRange.Columns.Count - 1
                    for c in range(1, last_col + 1):
                        if ws_com.Columns(c).Hidden:
                            ws_com.Columns(c).Hidden = False
                            sheet_cols += 1
                except Exception:
                    pass
                if sheet_rows or sheet_cols:
                    by_sheet[ws_com.Name] = {"rows": sheet_rows, "cols": sheet_cols}
                rows_unhidden += sheet_rows
                cols_unhidden += sheet_cols
            wb.Save()

    return {
        "saved_to": str(p),
        "rows_unhidden": rows_unhidden,
        "cols_unhidden": cols_unhidden,
        "by_sheet": by_sheet,
    }


# ============================================================
# Diff against snapshot
# ============================================================

def diff_totals(path: str | Path | None, snapshot: dict, alert_pct: float = 5.0) -> dict:
    current = compute_total(path)
    deltas = {}
    alerts = []
    for key in ("labor", "expenses", "grand"):
        prev = float(snapshot.get(key, 0))
        cur = float(current[key])
        delta = cur - prev
        deltas[key] = {"prev": prev, "current": cur, "delta": round(delta, 2)}
        if prev > 0 and abs(delta) / prev * 100 >= alert_pct:
            direction = "DECREASED" if delta < 0 else "INCREASED"
            alerts.append(
                f"{key.upper()} {direction} by {abs(delta/prev)*100:.1f}% "
                f"(${prev:,.0f} -> ${cur:,.0f})"
            )
    rounded_prev = _roundup_1k(float(snapshot.get("grand", 0)))
    rounded_cur = current["rounded_1k"]
    if rounded_cur != rounded_prev:
        alerts.append(f"Rounded total changed: ${rounded_prev:,} -> ${rounded_cur:,}")
    return {"deltas": deltas, "alerts": alerts, "current": current}


# ============================================================
# Fee-table payload for build.py
# ============================================================

def fee_table_payload(
    path: str | Path | None = None,
    title: str = "Fees for Professional Services",
    lead: Optional[str] = None,
) -> dict:
    """Build a fee_table section payload. Empty phases (no labor + no expense)
    are omitted — a 12-phase template used for a 7-phase project yields a
    7-phase payload."""
    full = read_workbook(path)
    phases = []
    for ph in full["phases"]:
        labor = int(round(ph["labor_subtotal"]))
        expense = int(round(ph["expense_subtotal"]))
        if labor == 0 and expense == 0:
            continue
        phases.append({"name": ph["name"], "labor": labor, "expenses": expense})
    payload = {"type": "fee_table", "title": title, "phases": phases}
    if lead:
        payload["lead"] = lead
    return payload
