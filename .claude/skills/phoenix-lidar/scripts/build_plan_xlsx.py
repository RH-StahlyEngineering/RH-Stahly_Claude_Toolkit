"""Build the processing-plan workbook from analyze.py outputs (matrix.csv + batches.json).
README + Processing Plan (boolean checkbox column, grouped by base batch) + No-Base sheet.

Usage: python build_plan_xlsx.py --out-dir <OUT> --out <OUT>/Processing_Plan.xlsx [--project "Name"]
"""
import argparse, os, csv, json
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

AR = lambda **k: Font(name="Arial", **k)
NAVY = "1F3864"; CBFMT = '[=1]"☑";[=0]"☐";;'


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--out-dir', required=True); ap.add_argument('--out', required=True)
    ap.add_argument('--project', default='Phoenix LiDAR')
    a = ap.parse_args()
    rows = {r['session']: r for r in csv.DictReader(open(os.path.join(a.out_dir, 'matrix.csv')))}
    batches = json.load(open(os.path.join(a.out_dir, 'batches.json')))
    # label duplicate files (a)/(b)
    seen = defaultdict(int); labelled = []
    fcount = defaultdict(int)
    for b in batches: fcount[b['file']] += 1
    for b in batches:
        suf = ""
        if fcount[b['file']] > 1:
            seen[b['file']] += 1; suf = f" ({chr(96+seen[b['file']])})"
        labelled.append((f"{b['base']} · {b['file']}{suf}", b))

    thin = Side(style="thin", color="BFBFBF"); border = Border(thin, thin, thin, thin)
    hdr = PatternFill("solid", fgColor=NAVY); grpfill = PatternFill("solid", fgColor="8EAADB")
    green = PatternFill("solid", fgColor="C6EFCE")
    ctr = Alignment(horizontal="center", vertical="center"); left = Alignment(horizontal="left", vertical="center")

    wb = Workbook(); ws = wb.active; ws.title = "Processing Plan"
    ws["A1"] = f"{a.project} — LiDAR PPK Processing Plan"; ws["A1"].font = AR(bold=True, size=14, color=NAVY)
    ws["A2"] = "One base file per batch, ≤3 missions each. Set Processed = TRUE (☑) when a mission's PPK is done."
    ws["A2"].font = AR(size=9, italic=True, color="595959"); ws.merge_cells("A2:J2")
    cols = ["Processed", "Batch", "Base", "Base file", "Session", "Type", "Scan start (UTC)", "Baseline (mi)", "Working RINEX", "Notes"]
    for c, n in enumerate(cols, 1):
        cell = ws.cell(4, c, n); cell.font = AR(bold=True, color="FFFFFF"); cell.fill = hdr
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = border
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False); ws.add_data_validation(dv)
    r = 5; datarows = []
    for label, b in labelled:
        ws.cell(r, 1, label).font = AR(bold=True, color="FFFFFF")
        for c in range(1, len(cols) + 1): ws.cell(r, c).fill = grpfill; ws.cell(r, c).border = border
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(cols)); r += 1
        for s in b['missions']:
            m = rows.get(s, {})
            mi = m.get('best_base_mi', '')
            note = ("baseline over rule" if (m.get('base_brackets') == 'YES' and m.get('has_working_rinex') == 'NO') else "")
            vals = [False, b['base'], m.get('best_base', ''), b['file'], s, m.get('type', ''),
                    m.get('scan_start_utc', ''), float(mi) if mi not in ('', None) else '', m.get('has_working_rinex', ''), note]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(r, c, v); cell.font = AR(size=10); cell.border = border
                cell.alignment = ctr if c in (1, 3, 5, 6, 7, 8, 9) else left
                if c == 8: cell.number_format = "0.0"
            ws.cell(r, 1).number_format = CBFMT; ws.cell(r, 1).font = AR(size=12); dv.add(ws.cell(r, 1))
            if note: ws.cell(r, 10).font = AR(size=10, bold=True, color="C00000")
            datarows.append(r); r += 1
    if datarows:
        f0, f1 = datarows[0], datarows[-1]
        ws.conditional_formatting.add(f"A{f0}:J{f1}", FormulaRule(formula=[f'$A{f0}=TRUE'], fill=green))
        ws.cell(f1 + 2, 2, "Processed:").font = AR(bold=True)
        ws.cell(f1 + 2, 3, f'=COUNTIF(A{f0}:A{f1},TRUE)').font = AR(bold=True, color="1F7244")
        ws.cell(f1 + 2, 4, f'of {len(datarows)} missions').font = AR(italic=True, color="595959")
    for i, wd in enumerate([11, 24, 7, 16, 17, 8, 18, 12, 14, 22], 1):
        ws.column_dimensions[chr(64 + i)].width = wd
    ws.freeze_panes = "A5"; ws.sheet_view.showGridLines = False

    # README
    rm = wb.create_sheet("README", 0); rm.sheet_view.showGridLines = False
    rm.column_dimensions['A'].width = 3; rm.column_dimensions['B'].width = 108; rr = [1]
    def line(t="", **k):
        c = rm.cell(rr[0], 2, t); c.font = AR(**k); c.alignment = Alignment(wrap_text=True, vertical="top"); rr[0] += 1
    line(f"{a.project} — LiDAR PPK Processing Plan", bold=True, size=15, color=NAVY); line()
    line("Base-station assignment and processing checklist for the valid LiDAR acquisitions.", italic=True, color="595959"); line()
    line("HOW TO USE", bold=True, color=NAVY)
    line("• 'Processed' is a boolean (TRUE/FALSE) shown as a checkbox ☑/☐. Set TRUE when a mission's PPK is done — the row turns green.")
    line("• For a clickable native checkbox: select column A in Excel → Insert → Checkbox (values are already boolean).")
    line("• The counter below the table tallies completed missions."); line()
    line("METHOD", bold=True, color=NAVY)
    line("• A base 'works' only if its RINEX obs file FULLY time-brackets the scan (no gap).")
    line("• Recommended base = shortest baseline among bracketing bases; batches = ≤3 missions sharing one base file.")
    line("• Baseline (mi) is the furthest-point (or average) distance from base to the mission trajectory."); line()
    line("SHEETS", bold=True, color=NAVY)
    line("• Processing Plan — the checklist, grouped into base batches.")
    line("• No Base — valid acquisitions no base covers in time+range.")

    # No-base sheet
    ws2 = wb.create_sheet("No Base"); ws2.sheet_view.showGridLines = False
    ws2["A1"] = "Valid acquisitions with no base in time+range"; ws2["A1"].font = AR(bold=True, size=12, color=NAVY)
    for c, n in enumerate(["Session", "Type", "Scan start (UTC)", "Brackets?", "Note"], 1):
        cell = ws2.cell(3, c, n); cell.font = AR(bold=True, color="FFFFFF"); cell.fill = hdr; cell.border = border
    rr2 = 4
    for s, m in sorted(rows.items()):
        if m.get('best_base_file'): continue
        for c, v in enumerate([s, m.get('type', ''), m.get('scan_start_utc', ''), m.get('base_brackets', ''), "no bracketing base"], 1):
            cell = ws2.cell(rr2, c, v); cell.font = AR(size=10); cell.border = border
        rr2 += 1
    for i, wd in enumerate([17, 8, 18, 10, 24], 1): ws2.column_dimensions[chr(64 + i)].width = wd

    wb.save(a.out)
    print(f"WROTE {a.out}  ({len(datarows)} missions in {len(labelled)} batches)")


if __name__ == '__main__':
    main()
